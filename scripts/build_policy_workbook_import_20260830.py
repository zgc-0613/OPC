import json
import re
import sys
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = Path(r"E:\个人发展\0-科研\OPC\opc-sci\调研表\人工智能OPC政策汇总表.xlsx")
OUT = ROOT / "deploy" / "sql" / "20260830_policy_workbook_classification.sql"

LABELS = ["算力技术", "财政激励", "场景开放", "创业生态", "金融资本", "制度治理", "人才培育"]
OLD_TYPE = {
    "算力技术": "computing_support",
    "财政激励": "funding_subsidy",
    "场景开放": "scenario_demand",
    "创业生态": "comprehensive",
    "金融资本": "investment",
    "制度治理": "governance_market",
    "人才培育": "talent_service",
}


def q(value):
    value = "" if value is None else str(value)
    return "'" + value.replace("\\", "\\\\").replace("'", "''") + "'"


def main():
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    ws = openpyxl.load_workbook(path, data_only=False).active
    rows = []
    for r in range(2, ws.max_row + 1):
        policy_id = ws.cell(r, 1).value
        if policy_id is None:
            continue
        categories = [LABELS[i] for i in range(7) if ws.cell(r, 4 + i).value == "是"]
        marker_text = " ".join(str(ws.cell(r, c).value or "") for c in range(1, 15))
        action = "keep"
        if "删除" in marker_text:
            action = "delete"
        elif "转为征求意见稿" in marker_text or "移入征求意见稿" in marker_text:
            action = "consultation"
        elif "移至标准规范" in marker_text:
            action = "standard"
        elif "更改为服务" in marker_text:
            action = "service"
        rows.append({"id": int(policy_id), "categories": categories, "action": action})

    ids = [row["id"] for row in rows]
    if len(rows) != 80 or len(set(ids)) != 80:
        raise RuntimeError(f"expected 80 unique rows, got {len(rows)}")
    classified = [row for row in rows if row["action"] == "keep" and row["categories"]]
    if len(classified) != 71:
        raise RuntimeError(f"expected 71 classified retained rows, got {len(classified)}")
    delete_ids = [row["id"] for row in rows if row["action"] == "delete"]
    if delete_ids != [14, 21, 33, 60]:
        raise RuntimeError(f"unexpected delete set: {delete_ids}")

    target_ids = ",".join(map(str, ids))
    delete_list = ",".join(map(str, delete_ids))
    category_cases = []
    type_cases = []
    relation_rows = []
    for row in rows:
        if row["action"] != "keep" or not row["categories"]:
            continue
        categories = row["categories"]
        category_cases.append(f"WHEN {row['id']} THEN {q('，'.join(categories))}")
        relation_rows.extend((row["id"], category) for category in categories)
        policy_type = "comprehensive" if len(categories) >= 3 else OLD_TYPE[categories[0]]
        type_cases.append(f"WHEN {row['id']} THEN {q(policy_type)}")

    backups = f"""CREATE TABLE IF NOT EXISTS policy_workbook_backup_20260830 LIKE policies;
INSERT IGNORE INTO policy_workbook_backup_20260830
SELECT * FROM policies WHERE id IN ({target_ids});
"""
    updates = f"""UPDATE policies SET
  policy_type = CASE id
    {' '.join(type_cases)}
    ELSE policy_type END,
  tags = CASE id
    {' '.join(category_cases)}
    ELSE tags END,
  reviewer = CASE id
    {' '.join(f"WHEN {row['id']} THEN {q('manual-classification-workbook-20260830')}" for row in rows if row['action'] == 'keep' and row['categories'])}
    ELSE reviewer END
WHERE id IN ({target_ids});

INSERT INTO tags (name, tag_type, is_industry, sort_order) VALUES
{', '.join(f"({q(label)}, 'policy', 0, {index})" for index, label in enumerate(LABELS, 1))}
ON DUPLICATE KEY UPDATE sort_order=VALUES(sort_order), is_industry=0;
DELETE FROM policy_tags WHERE policy_id IN ({target_ids});
INSERT IGNORE INTO policy_tags (policy_id, tag_id)
SELECT relation.policy_id, dictionary.id
FROM (
  {' UNION ALL '.join(f"SELECT {policy_id} AS policy_id, {q(category)} AS tag_name" for policy_id, category in relation_rows)}
) relation
JOIN tags dictionary ON dictionary.name=relation.tag_name AND dictionary.tag_type='policy';

UPDATE policies SET material_nature='consultation_draft', status='consultation'
WHERE id IN (40,121);
UPDATE policies SET material_nature='standard_reference'
WHERE id IN (82,124);
UPDATE policies SET material_nature='official_platform_service'
WHERE id IN (92);
UPDATE policies SET tags='', policy_type='other', reviewer='material-adjustment-workbook-20260830'
WHERE id IN (40,82,92,121,124);

DELETE FROM policy_industry_tags WHERE policy_id IN ({delete_list});
DELETE FROM policy_tags WHERE policy_id IN ({delete_list});
DELETE FROM policies WHERE id IN ({delete_list});

UPDATE policies SET original_url='http://www.cixi.gov.cn/col/col1229039850/art/2026/art_43506befefd5420085f1c54e6fa78a9f.html'
WHERE id=20;
UPDATE sources s JOIN policies p ON p.source_id=s.id
SET s.url='http://www.cixi.gov.cn/col/col1229039850/art/2026/art_43506befefd5420085f1c54e6fa78a9f.html'
WHERE p.id=20;
UPDATE policies SET original_url='https://jsip.jiangsu.gov.cn/attach/0/511551c6565f48e0bbe6c9acf7ae493b.docx'
WHERE id=82;
UPDATE sources s JOIN policies p ON p.source_id=s.id
SET s.url='https://jsip.jiangsu.gov.cn/attach/0/511551c6565f48e0bbe6c9acf7ae493b.docx'
WHERE p.id=82;
"""
    checks = f"""SELECT CONCAT('target_remaining=',COUNT(*)) FROM policies WHERE id IN ({target_ids});
SELECT CONCAT('deleted_remaining=',COUNT(*)) FROM policies WHERE id IN ({delete_list});
SELECT CONCAT('classified=',COUNT(*)) FROM policies WHERE id IN ({target_ids}) AND reviewer='manual-classification-workbook-20260830' AND tags IS NOT NULL AND tags<>'';
SELECT CONCAT('empty_classification=',COUNT(*)) FROM policies WHERE id IN ({target_ids}) AND id NOT IN ({delete_list}) AND reviewer='manual-classification-workbook-20260830' AND (tags IS NULL OR tags='');
SELECT material_nature,COUNT(*) FROM policies WHERE id IN ({target_ids}) GROUP BY material_nature ORDER BY material_nature;
"""
    sql = "START TRANSACTION;\n" + backups + updates + checks + "COMMIT;\n"
    OUT.write_text(sql, encoding="utf-8")
    print(json.dumps({"rows": len(rows), "classified": len(classified), "delete": delete_ids, "output": str(OUT)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
