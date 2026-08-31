#!/usr/bin/env python
"""Prepare the reviewed V4 case migration without connecting to production.

The script treats the uploaded V4 workbook and the paper snapshot as immutable
inputs. It validates the approved 167-case result, writes a machine-readable
plan, and emits a one-time MySQL migration with backups and assertions.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import unicodedata
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any
from urllib.parse import parse_qsl, urlencode, urlsplit, urlunsplit

import openpyxl


EXPECTED_EXCEL_SHA256 = "21191265952269AA90F1DFCBAC3EA1365C077AD0D209E079DD5D408B4B565997"
EXPECTED_SNAPSHOT_SHA256 = "21DFBDEDB8118AF6614F1690C226C54DFEC9BFB40A79005EF60C63BDD9E4C594"
EXPECTED_INPUT_CASES = 177
EXPECTED_FINAL_CASES = 167
EXPECTED_PRODUCTION_CASES = 154
EXPECTED_UPDATES = 148
EXPECTED_INSERTS = 19
EXPECTED_DELETES = 6
EXPECTED_DELETE_IDS = {124, 125, 127, 146, 148, 152}
MAJOR_CATEGORIES = {
    "内容创作": ("短剧动漫", "图文音频", "游戏虚拟", "文化IP", "文旅体验"),
    "商业增长": ("电商运营", "跨境电商", "广告种草", "智能零售"),
    "软件工具": ("应用开发", "办公智能体", "数据服务", "语音交互"),
    "教育人才": ("学习产品", "培训课程", "就业服务"),
    "产业应用": ("工业能源", "城市空间", "智能硬件", "健康养老", "农业生产", "农产品品牌营销", "乡村综合服务"),
    "创业支撑": ("活动赛事", "算力服务"),
}
PLACEHOLDERS = {"43", "44"}


@dataclass
class CaseRecord:
    excel_row: int
    article_title: str
    title: str
    original_url: str
    province: str | None
    city: str | None
    district: str | None
    actor_name: str
    student_startup: str | None
    category: str
    subcategory: str
    publish_date: str | None
    summary: str
    ai_tools: str | None
    outcome: str | None
    business_model: str = ""
    tags: str | None = None
    production_id: int | None = None
    source_id: int | None = None
    replace_source: bool = False


@dataclass(frozen=True)
class DuplicateDecision:
    keep_row: int
    drop_row: int
    category: str
    subcategory: str


DUPLICATE_DECISIONS = (
    DuplicateDecision(29, 118, "产业应用", "智能硬件"),
    DuplicateDecision(61, 143, "商业增长", "跨境电商"),
    DuplicateDecision(62, 112, "产业应用", "工业能源"),
    DuplicateDecision(63, 113, "商业增长", "广告种草"),
    DuplicateDecision(64, 114, "教育人才", "学习产品"),
    DuplicateDecision(37, 144, "软件工具", "应用开发"),
    DuplicateDecision(100, 126, "内容创作", "短剧动漫"),
    DuplicateDecision(18, 145, "产业应用", "工业能源"),
    DuplicateDecision(91, 173, "商业增长", "电商运营"),
    DuplicateDecision(180, 44, "内容创作", "短剧动漫"),
)

# These are continuity mappings, not duplicate additions. Keeping the existing
# IDs preserves visit, review and AI-run history while replacing the content.
REPLACEMENT_IDS = {97: 97, 101: 101, 180: 44}

PUBLISHERS = {
    "suda.edu.cn": "苏州大学",
    "cnr.cn": "央广网",
    "rzw.com.cn": "日照网",
    "whsw.cn": "武汉商学院",
    "changsha.gov.cn": "长沙市人民政府",
    "ncss.cn": "国家大学生就业服务平台",
    "eyesnews.cn": "贵州日报天眼新闻",
    "zjbti.net.cn": "浙江工商职业技术学院",
    "qdxin.cn": "信网",
    "chinanews.com.cn": "中国新闻网",
    "xinhuanet.com": "新华网",
    "china.org.cn": "中国网",
}


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", type=Path, required=True)
    parser.add_argument("--snapshot", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument("--batch", default="casev4_20260817")
    parser.add_argument("--database", default="opc_platform")
    parser.add_argument("--accessed-at", default=date.today().isoformat())
    parser.add_argument("--allow-different-hash", action="store_true")
    args = parser.parse_args()

    validate_hash(args.excel, EXPECTED_EXCEL_SHA256, args.allow_different_hash)
    validate_hash(args.snapshot, EXPECTED_SNAPSHOT_SHA256, args.allow_different_hash)
    records = read_excel(args.excel)
    snapshot = read_snapshot(args.snapshot)
    final_records, duplicate_report = apply_duplicate_decisions(records)
    migration = match_production(final_records, snapshot)
    validate_plan(records, final_records, migration)

    args.output_dir.mkdir(parents=True, exist_ok=True)
    plan_path = args.output_dir / "case-v4-plan.json"
    sql_path = args.output_dir / "case-v4-migration.sql"
    rollback_path = args.output_dir / "case-v4-rollback.sql"
    plan = build_plan(args, records, final_records, duplicate_report, migration)
    plan_path.write_text(json.dumps(plan, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    sql_path.write_text(generate_sql(args, final_records, snapshot, migration), encoding="utf-8")
    rollback_path.write_text(generate_rollback_sql(args), encoding="utf-8")
    print(json.dumps(plan["summary"], ensure_ascii=False, indent=2))
    print(f"Plan: {plan_path}")
    print(f"SQL:  {sql_path}")
    print(f"Rollback: {rollback_path}")


def validate_hash(path: Path, expected: str, allow_different: bool) -> None:
    actual = hashlib.sha256(path.read_bytes()).hexdigest().upper()
    if actual != expected and not allow_different:
        raise RuntimeError(f"Unexpected input hash for {path}: {actual}; expected {expected}")


def decode_text(value: Any) -> Any:
    # The corrected V4 workbook is already Unicode. Re-decoding Latin-1-safe
    # text as GBK corrupts valid names such as "ChenYe·Wordflow" because the
    # middle dot and following ASCII letter form a valid but unrelated GBK pair.
    return value


def clean(value: Any, *, placeholder_null: bool = True) -> str | None:
    value = decode_text(value)
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = re.sub(r"\s+", " ", str(value).strip())
    if not text or (placeholder_null and text in PLACEHOLDERS):
        return None
    return text


def parse_date(value: Any) -> str | None:
    direct = clean(value)
    if not direct:
        return None
    match = re.search(r"(\d{4})[-/.年](\d{1,2})[-/.月](\d{1,2})", direct)
    if not match:
        return None
    try:
        return date(*(int(part) for part in match.groups())).isoformat()
    except ValueError:
        return None


def read_excel(path: Path) -> list[CaseRecord]:
    workbook = openpyxl.load_workbook(path, data_only=True)
    sheet = workbook.worksheets[0]
    headers = [clean(sheet.cell(2, column).value, placeholder_null=False) for column in range(1, 15)]
    expected = [
        "文章名", "命名", "案例来源（链接）", "关联地区-省", "关联地区-市", "关联地区-区",
        "主体名", "是否为大学生创新创业", "大类", "小类", "发布日期", "100-300字摘要",
        "用到的AI工具或能力", "结果（效果、成果、数据）",
    ]
    if headers != expected:
        raise RuntimeError(f"Unexpected case headers: {headers}")
    records: list[CaseRecord] = []
    for row in range(3, sheet.max_row + 1):
        values = [sheet.cell(row, column).value for column in range(1, 15)]
        title = clean(values[1])
        source_url = clean(values[2])
        if not title or not source_url:
            continue
        record = CaseRecord(
            excel_row=row,
            article_title=required(values[0], row, "文章名"),
            title=title,
            original_url=source_url,
            province=clean(values[3]),
            city=clean(values[4]),
            district=clean(values[5]),
            actor_name=required(values[6], row, "主体名"),
            student_startup=clean(values[7]),
            category=required(values[8], row, "大类"),
            subcategory=required(values[9], row, "小类"),
            publish_date=parse_date(values[10]),
            summary=required(values[11], row, "摘要"),
            ai_tools=clean(values[12]),
            outcome=clean(values[13]),
        )
        validate_taxonomy(record)
        records.append(record)
    if len(records) != EXPECTED_INPUT_CASES:
        raise RuntimeError(f"Expected {EXPECTED_INPUT_CASES} Excel cases, found {len(records)}")
    return records


def required(value: Any, row: int, field: str) -> str:
    result = clean(value)
    if not result:
        raise RuntimeError(f"Excel row {row}: missing {field}")
    return result


def validate_taxonomy(record: CaseRecord) -> None:
    if record.category not in MAJOR_CATEGORIES or record.subcategory not in MAJOR_CATEGORIES[record.category]:
        raise RuntimeError(
            f"Excel row {record.excel_row}: invalid taxonomy {record.category}/{record.subcategory}"
        )


def apply_duplicate_decisions(
    records: list[CaseRecord],
) -> tuple[list[CaseRecord], list[dict[str, Any]]]:
    by_row = {record.excel_row: record for record in records}
    dropped: set[int] = set()
    report: list[dict[str, Any]] = []
    for decision in DUPLICATE_DECISIONS:
        keep = by_row[decision.keep_row]
        drop = by_row[decision.drop_row]
        if normalized_url(keep.original_url) == normalized_url(drop.original_url):
            raise RuntimeError(f"Rows {keep.excel_row}/{drop.excel_row} are not cross-report duplicates")
        keep.category = decision.category
        keep.subcategory = decision.subcategory
        if not keep.ai_tools:
            keep.ai_tools = drop.ai_tools
        if not keep.outcome:
            keep.outcome = drop.outcome
        validate_taxonomy(keep)
        dropped.add(drop.excel_row)
        report.append({
            "keep_row": keep.excel_row,
            "keep_title": keep.title,
            "keep_url": keep.original_url,
            "drop_row": drop.excel_row,
            "drop_title": drop.title,
            "drop_url": drop.original_url,
            "final_category": keep.category,
            "final_subcategory": keep.subcategory,
        })
    final = [record for record in records if record.excel_row not in dropped]
    for record in final:
        record.business_model = build_business_model(record)
        record.tags = build_tags(record)
    return final, report


def build_business_model(record: CaseRecord) -> str:
    region = " / ".join(part for part in (record.province, record.city, record.district) if part)
    lines = [f"地区：{region}" if region else "地区：原文未明确"]
    lines.append(f"案例类型：{record.category} / {record.subcategory}")
    if record.student_startup:
        lines.append(f"是否为大学生创新创业：{record.student_startup}")
    if record.publish_date:
        lines.append(f"发布日期：{record.publish_date}")
    return "\n".join(lines)


def build_tags(record: CaseRecord) -> str | None:
    tags: list[str] = []
    if record.student_startup == "是":
        tags.append("大学生创新创业")
    if record.ai_tools:
        for part in re.split(r"[;；、,，]", record.ai_tools):
            part = part.strip()
            if 1 < len(part) <= 30 and part not in tags:
                tags.append(part)
            if len(tags) >= 8:
                break
    return ",".join(tags) or None


def read_snapshot(path: Path) -> dict[str, Any]:
    workbook = openpyxl.load_workbook(path, data_only=True)
    case_sheet = workbook["cases_full"]
    source_sheet = workbook["sources_full"]
    region_sheet = workbook["regions"]
    cases = sheet_dicts(case_sheet)
    sources = sheet_dicts(source_sheet)
    regions = sheet_dicts(region_sheet)
    if len(cases) != EXPECTED_PRODUCTION_CASES:
        raise RuntimeError(f"Expected {EXPECTED_PRODUCTION_CASES} production cases, found {len(cases)}")
    return {
        "cases": cases,
        "sources": sources,
        "regions": regions,
        "source_by_url": {normalized_url(row.get("URL")): row for row in sources if row.get("URL")},
        "region_by_name": {str(row.get("Name")): int(row["ID"]) for row in regions if row.get("Name")},
    }


def sheet_dicts(sheet: Any) -> list[dict[str, Any]]:
    headers = [str(cell.value) for cell in sheet[1]]
    rows: list[dict[str, Any]] = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        if values[0] is None:
            continue
        rows.append({header: values[index] for index, header in enumerate(headers)})
    return rows


def match_production(records: list[CaseRecord], snapshot: dict[str, Any]) -> dict[str, Any]:
    summary_index: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in snapshot["cases"]:
        summary_index[normalized_text(row.get("Summary"))].append(row)
    used_ids: set[int] = set()
    updates: list[CaseRecord] = []
    inserts: list[CaseRecord] = []
    for record in records:
        replacement_id = REPLACEMENT_IDS.get(record.excel_row)
        if replacement_id is not None:
            record.production_id = replacement_id
            record.replace_source = True
        else:
            matches = summary_index.get(normalized_text(record.summary), [])
            if len(matches) == 1:
                record.production_id = int(matches[0]["ID"])
                record.source_id = int(matches[0]["Source ID"])
        if record.production_id is None:
            inserts.append(record)
        else:
            if record.production_id in used_ids:
                raise RuntimeError(f"Production case ID reused: {record.production_id}")
            used_ids.add(record.production_id)
            updates.append(record)
    production_ids = {int(row["ID"]) for row in snapshot["cases"]}
    delete_ids = sorted(production_ids - used_ids)
    return {"updates": updates, "inserts": inserts, "delete_ids": delete_ids}


def validate_plan(
    input_records: list[CaseRecord], final_records: list[CaseRecord], migration: dict[str, Any]
) -> None:
    if len(input_records) != EXPECTED_INPUT_CASES or len(final_records) != EXPECTED_FINAL_CASES:
        raise RuntimeError("Unexpected case count after deduplication")
    if len(migration["updates"]) != EXPECTED_UPDATES:
        raise RuntimeError(f"Expected {EXPECTED_UPDATES} updates, found {len(migration['updates'])}")
    if len(migration["inserts"]) != EXPECTED_INSERTS:
        raise RuntimeError(f"Expected {EXPECTED_INSERTS} inserts, found {len(migration['inserts'])}")
    if len(migration["delete_ids"]) != EXPECTED_DELETES:
        raise RuntimeError(f"Expected {EXPECTED_DELETES} deletes, found {len(migration['delete_ids'])}")
    if set(migration["delete_ids"]) != EXPECTED_DELETE_IDS:
        raise RuntimeError(f"Unexpected delete IDs: {migration['delete_ids']}")
    if len({record.title for record in final_records}) != len(final_records):
        raise RuntimeError("Duplicate display names remain after deduplication")
    if any(value in PLACEHOLDERS for record in final_records for value in asdict(record).values()):
        raise RuntimeError("43/44 placeholder survived normalization")
    counts = Counter(record.category for record in final_records)
    if set(counts) != set(MAJOR_CATEGORIES):
        raise RuntimeError(f"Unexpected major-category set: {counts}")


def build_plan(
    args: argparse.Namespace,
    input_records: list[CaseRecord],
    final_records: list[CaseRecord],
    duplicate_report: list[dict[str, Any]],
    migration: dict[str, Any],
) -> dict[str, Any]:
    return {
        "summary": {
            "input_cases": len(input_records),
            "duplicate_groups": len(duplicate_report),
            "final_cases": len(final_records),
            "updates": len(migration["updates"]),
            "inserts": len(migration["inserts"]),
            "deletes": len(migration["delete_ids"]),
            "delete_ids": migration["delete_ids"],
            "major_categories": Counter(record.category for record in final_records),
            "null_ai_tools": sum(record.ai_tools is None for record in final_records),
            "null_outcomes": sum(record.outcome is None for record in final_records),
        },
        "inputs": {
            "excel": str(args.excel.resolve()),
            "excel_sha256": sha256(args.excel),
            "snapshot": str(args.snapshot.resolve()),
            "snapshot_sha256": sha256(args.snapshot),
        },
        "duplicates": duplicate_report,
        "updates": [record_plan(record) for record in migration["updates"]],
        "inserts": [record_plan(record) for record in migration["inserts"]],
        "delete_ids": migration["delete_ids"],
    }


def record_plan(record: CaseRecord) -> dict[str, Any]:
    return {
        "excel_row": record.excel_row,
        "production_id": record.production_id,
        "title": record.title,
        "article_title": record.article_title,
        "category": record.category,
        "subcategory": record.subcategory,
        "source_url": record.original_url,
        "replace_source": record.replace_source,
    }


def generate_sql(
    args: argparse.Namespace,
    records: list[CaseRecord],
    snapshot: dict[str, Any],
    migration: dict[str, Any],
) -> str:
    backup_suffix = re.sub(r"[^0-9a-zA-Z_]", "_", args.batch)
    database = mysql_identifier(args.database)
    lines = [
        "-- Generated by scripts/prepare_case_v4_migration.py",
        f"-- Excel SHA-256: {sha256(args.excel)}",
        f"-- Snapshot SHA-256: {sha256(args.snapshot)}",
        "-- One-time migration. Existing backup-table names intentionally make reruns fail closed.",
        "SET NAMES utf8mb4;",
        f"USE {database};",
        "",
        assertion_procedure(EXPECTED_PRODUCTION_CASES, EXPECTED_FINAL_CASES),
        "CALL assert_case_v4_preflight();",
        "",
        ddl_if_missing("article_title", "ALTER TABLE case_items ADD COLUMN article_title VARCHAR(500) NULL COMMENT 'Original source article title' AFTER title"),
        ddl_if_missing("subcategory", "ALTER TABLE case_items ADD COLUMN subcategory VARCHAR(50) NULL COMMENT 'Twenty-five-subcategory case taxonomy' AFTER category"),
        index_if_missing("idx_case_items_subcategory", "ALTER TABLE case_items ADD INDEX idx_case_items_subcategory (subcategory)"),
        "",
        f"CREATE TABLE backup_case_items_{backup_suffix} LIKE case_items;",
        f"INSERT INTO backup_case_items_{backup_suffix} SELECT * FROM case_items;",
        f"CREATE TABLE backup_sources_{backup_suffix} LIKE sources;",
        f"INSERT INTO backup_sources_{backup_suffix} SELECT * FROM sources;",
        f"CREATE TABLE backup_tags_{backup_suffix} LIKE tags;",
        f"INSERT INTO backup_tags_{backup_suffix} SELECT * FROM tags;",
        f"CREATE TABLE backup_case_tags_{backup_suffix} LIKE case_tags;",
        f"INSERT INTO backup_case_tags_{backup_suffix} SELECT * FROM case_tags;",
        optional_statement("visit_logs", f"CREATE TABLE backup_visit_logs_{backup_suffix} LIKE visit_logs"),
        optional_statement("visit_logs", f"INSERT INTO backup_visit_logs_{backup_suffix} SELECT * FROM visit_logs"),
        optional_statement(
            "ai_analysis_runs",
            f"CREATE TABLE backup_ai_analysis_runs_{backup_suffix} AS SELECT id, case_id FROM ai_analysis_runs",
        ),
        optional_statement("evidence_reviews", f"CREATE TABLE backup_evidence_reviews_{backup_suffix} LIKE evidence_reviews"),
        optional_statement("evidence_reviews", f"INSERT INTO backup_evidence_reviews_{backup_suffix} SELECT * FROM evidence_reviews"),
        "START TRANSACTION;",
        "CREATE TEMPORARY TABLE tmp_case_v4_ids (excel_row INT PRIMARY KEY, case_id BIGINT NOT NULL UNIQUE);",
        "",
    ]

    snapshot_case_by_id = {int(row["ID"]): row for row in snapshot["cases"]}
    for record in records:
        if record.production_id is not None:
            if record.replace_source:
                lines.extend(ensure_source_sql(record, args.accessed_at, snapshot))
                source_expression = "@case_v4_source_id"
                evidence_status_assignment = ", ai_evidence_status='legacy_unverified'"
            else:
                source_expression = str(int(record.source_id or snapshot_case_by_id[record.production_id]["Source ID"]))
                evidence_status_assignment = ""
            region_expression = region_sql(record.province)
            lines.extend([
                f"-- Excel row {record.excel_row}: update case ID {record.production_id}",
                "UPDATE case_items SET",
                f"  title={sql(record.title)}, article_title={sql(record.article_title)},",
                f"  region_id={region_expression}, category={sql(record.category)}, subcategory={sql(record.subcategory)},",
                f"  actor_name={sql(record.actor_name)}, source_id={source_expression}, summary={sql(record.summary)},",
                f"  business_model={sql(record.business_model)}, ai_tools={sql(record.ai_tools)}, outcome={sql(record.outcome)},",
                f"  tags={sql(record.tags)}, original_url={sql(record.original_url)}, status='published',",
                f"  reviewer='case-v4-migration', evidence_revision=COALESCE(evidence_revision,0)+1{evidence_status_assignment}",
                f"WHERE id={record.production_id};",
                f"INSERT INTO tmp_case_v4_ids VALUES ({record.excel_row}, {record.production_id});",
            ])
        else:
            lines.extend(ensure_source_sql(record, args.accessed_at, snapshot))
            lines.extend([
                f"-- Excel row {record.excel_row}: insert case",
                "INSERT INTO case_items (",
                "  title, article_title, region_id, category, subcategory, actor_name, source_id, summary,",
                "  business_model, ai_tools, outcome, tags, original_url, local_file, accessed_at, status, reviewer,",
                "  ai_evidence_status, evidence_revision",
                ") VALUES (",
                f"  {sql(record.title)}, {sql(record.article_title)}, {region_sql(record.province)}, {sql(record.category)},",
                f"  {sql(record.subcategory)}, {sql(record.actor_name)}, @case_v4_source_id, {sql(record.summary)},",
                f"  {sql(record.business_model)}, {sql(record.ai_tools)}, {sql(record.outcome)}, {sql(record.tags)},",
                f"  {sql(record.original_url)}, NULL, {sql(args.accessed_at)}, 'published', 'case-v4-migration',",
                "  'legacy_unverified', 0",
                ");",
                "SET @case_v4_case_id := LAST_INSERT_ID();",
                f"INSERT INTO tmp_case_v4_ids VALUES ({record.excel_row}, @case_v4_case_id);",
            ])
        lines.extend(rebuild_tags_sql(record))
        lines.append("")

    merge_pairs = [(152, 29), (127, 61), (148, 63), (124, 37), (146, 100), (125, 18)]
    for old_id, keep_row in merge_pairs:
        lines.extend([
            f"SET @case_v4_keep_id := (SELECT case_id FROM tmp_case_v4_ids WHERE excel_row={keep_row});",
            "INSERT IGNORE INTO case_tags (case_id, tag_id, created_at)",
            f"SELECT @case_v4_keep_id, tag_id, created_at FROM case_tags WHERE case_id={old_id};",
            optional_statement("visit_logs", f"UPDATE visit_logs SET target_id=@case_v4_keep_id WHERE target_type='case' AND target_id={old_id}"),
            optional_statement("ai_analysis_runs", f"UPDATE ai_analysis_runs SET case_id=@case_v4_keep_id WHERE case_id={old_id}"),
            optional_statement("evidence_reviews", f"UPDATE evidence_reviews SET item_id=@case_v4_keep_id WHERE item_type='case' AND item_id={old_id}"),
        ])
    delete_csv = ",".join(str(value) for value in migration["delete_ids"])
    lines.extend([
        f"DELETE FROM case_tags WHERE case_id IN ({delete_csv});",
        f"DELETE FROM case_items WHERE id IN ({delete_csv});",
        "CALL assert_case_v4_postflight();",
        "COMMIT;",
        "DROP PROCEDURE assert_case_v4_preflight;",
        "DROP PROCEDURE assert_case_v4_postflight;",
        "",
        "SELECT COUNT(*) AS final_case_count FROM case_items;",
        "SELECT category, COUNT(*) AS case_count FROM case_items GROUP BY category ORDER BY category;",
        "SELECT COUNT(*) AS invalid_taxonomy_count FROM case_items",
        "WHERE category NOT IN ('内容创作','商业增长','软件工具','教育人才','产业应用','创业支撑')",
        "   OR subcategory IS NULL OR TRIM(subcategory)='';",
        "",
    ])
    return "\n".join(lines)


def generate_rollback_sql(args: argparse.Namespace) -> str:
    backup_suffix = re.sub(r"[^0-9a-zA-Z_]", "_", args.batch)
    database = mysql_identifier(args.database)
    lines = [
        "-- Generated by scripts/prepare_case_v4_migration.py",
        "-- Emergency rollback to the exact pre-migration backup. Stop application writes first.",
        "SET NAMES utf8mb4;",
        f"USE {database};",
        "START TRANSACTION;",
        "DELETE FROM case_tags;",
        f"INSERT INTO case_tags SELECT * FROM backup_case_tags_{backup_suffix};",
        "DELETE FROM case_items;",
        f"INSERT INTO case_items SELECT * FROM backup_case_items_{backup_suffix};",
        f"DELETE current_source FROM sources current_source LEFT JOIN backup_sources_{backup_suffix} backup ON backup.id=current_source.id WHERE backup.id IS NULL;",
        f"DELETE current_tag FROM tags current_tag LEFT JOIN backup_tags_{backup_suffix} backup ON backup.id=current_tag.id WHERE backup.id IS NULL;",
        f"UPDATE tags current_tag JOIN backup_tags_{backup_suffix} backup ON backup.id=current_tag.id SET current_tag.is_industry=backup.is_industry;",
        restore_optional_table("visit_logs", f"backup_visit_logs_{backup_suffix}"),
        optional_statement(
            f"backup_ai_analysis_runs_{backup_suffix}",
            f"UPDATE ai_analysis_runs current_run JOIN backup_ai_analysis_runs_{backup_suffix} backup "
            "ON backup.id=current_run.id SET current_run.case_id=backup.case_id",
        ),
        restore_optional_table("evidence_reviews", f"backup_evidence_reviews_{backup_suffix}"),
        "COMMIT;",
        "SELECT COUNT(*) AS restored_case_count FROM case_items;",
    ]
    return "\n".join(lines) + "\n"


def ensure_source_sql(record: CaseRecord, accessed_at: str, snapshot: dict[str, Any]) -> list[str]:
    existing = snapshot["source_by_url"].get(normalized_url(record.original_url))
    if existing:
        return [f"SET @case_v4_source_id := {int(existing['ID'])};"]
    publisher = publisher_for(record.original_url)
    source_type = "government_site" if ".gov.cn" in record.original_url.lower() else "news"
    notes = f"case-v4-migration; excel_row={record.excel_row}"
    return [
        "INSERT INTO sources (title, source_type, publisher, url, local_file, accessed_at, notes, status, ai_evidence_status, evidence_revision)",
        f"SELECT {sql(record.article_title)}, {sql(source_type)}, {sql(publisher)}, {sql(record.original_url)}, NULL,",
        f"       {sql(accessed_at)}, {sql(notes)}, 'published', 'legacy_unverified', 0",
        f"WHERE NOT EXISTS (SELECT 1 FROM sources WHERE url={sql(record.original_url)});",
        f"SET @case_v4_source_id := (SELECT id FROM sources WHERE url={sql(record.original_url)} ORDER BY id LIMIT 1);",
    ]


def rebuild_tags_sql(record: CaseRecord) -> list[str]:
    values = [(record.category, True), (record.subcategory, False)]
    for tag in (record.tags or "").split(","):
        if tag.strip() and tag.strip() not in {record.category, record.subcategory}:
            values.append((tag.strip(), False))
    lines = [
        f"SET @case_v4_case_id := (SELECT case_id FROM tmp_case_v4_ids WHERE excel_row={record.excel_row});",
        "DELETE FROM case_tags WHERE case_id=@case_v4_case_id;",
    ]
    for name, industry in values:
        lines.extend([
            "INSERT INTO tags (name, tag_type, is_industry, sort_order)",
            f"VALUES ({sql(name)}, 'case', {1 if industry else 0}, 0)",
            f"ON DUPLICATE KEY UPDATE id=LAST_INSERT_ID(id), is_industry=GREATEST(is_industry,{1 if industry else 0});",
            "SET @case_v4_tag_id := LAST_INSERT_ID();",
            "INSERT IGNORE INTO case_tags (case_id, tag_id) VALUES (@case_v4_case_id, @case_v4_tag_id);",
        ])
    return lines


def ddl_if_missing(column: str, ddl: str) -> str:
    return "\n".join([
        f"SET @case_v4_ddl := IF((SELECT COUNT(*) FROM information_schema.columns WHERE table_schema=DATABASE() AND table_name='case_items' AND column_name={sql(column)})=0, {sql(ddl)}, 'SELECT 1');",
        "PREPARE case_v4_stmt FROM @case_v4_ddl; EXECUTE case_v4_stmt; DEALLOCATE PREPARE case_v4_stmt;",
    ])


def index_if_missing(index: str, ddl: str) -> str:
    return "\n".join([
        f"SET @case_v4_ddl := IF((SELECT COUNT(*) FROM information_schema.statistics WHERE table_schema=DATABASE() AND table_name='case_items' AND index_name={sql(index)})=0, {sql(ddl)}, 'SELECT 1');",
        "PREPARE case_v4_stmt FROM @case_v4_ddl; EXECUTE case_v4_stmt; DEALLOCATE PREPARE case_v4_stmt;",
    ])


def optional_statement(table: str, statement: str) -> str:
    return "\n".join([
        f"SET @case_v4_optional := IF((SELECT COUNT(*) FROM information_schema.tables WHERE table_schema=DATABASE() AND table_name={sql(table)})=1, {sql(statement)}, 'SELECT 1');",
        "PREPARE case_v4_optional_stmt FROM @case_v4_optional; EXECUTE case_v4_optional_stmt; DEALLOCATE PREPARE case_v4_optional_stmt;",
    ])


def restore_optional_table(table: str, backup_table: str) -> str:
    delete_sql = optional_statement(backup_table, f"DELETE FROM {table}")
    restore_sql = "\n".join([
        f"SET @case_v4_optional := IF((SELECT COUNT(*) FROM information_schema.tables WHERE table_schema=DATABASE() AND table_name={sql(backup_table)})=1, {sql(f'INSERT INTO {table} SELECT * FROM {backup_table}')}, 'SELECT 1');",
        "PREPARE case_v4_optional_stmt FROM @case_v4_optional; EXECUTE case_v4_optional_stmt; DEALLOCATE PREPARE case_v4_optional_stmt;",
    ])
    return delete_sql + "\n" + restore_sql


def assertion_procedure(pre_count: int, post_count: int) -> str:
    return f"""DELIMITER $$
CREATE PROCEDURE assert_case_v4_preflight()
BEGIN
  IF (SELECT COUNT(*) FROM case_items) <> {pre_count} THEN
    SIGNAL SQLSTATE '45000' SET MESSAGE_TEXT='case-v4 preflight failed: expected {pre_count} cases';
  END IF;
END$$
CREATE PROCEDURE assert_case_v4_postflight()
BEGIN
  IF (SELECT COUNT(*) FROM case_items) <> {post_count} THEN
    SIGNAL SQLSTATE '45000' SET MESSAGE_TEXT='case-v4 postflight failed: expected {post_count} cases';
  END IF;
  IF EXISTS (SELECT 1 FROM case_items WHERE article_title IS NULL OR TRIM(article_title)=''
             OR subcategory IS NULL OR TRIM(subcategory)='') THEN
    SIGNAL SQLSTATE '45000' SET MESSAGE_TEXT='case-v4 postflight failed: missing article title or subcategory';
  END IF;
END$$
DELIMITER ;"""


def publisher_for(url: str) -> str:
    host = urlsplit(url).netloc.lower()
    for suffix, publisher in PUBLISHERS.items():
        if host == suffix or host.endswith("." + suffix):
            return publisher
    return host


def region_sql(province: str | None) -> str:
    if not province:
        return "NULL"
    return f"(SELECT id FROM regions WHERE name={sql(province)} ORDER BY id LIMIT 1)"


def normalized_text(value: Any) -> str:
    text = unicodedata.normalize("NFKC", str(value or "")).lower()
    return re.sub(r"[^0-9a-z\u4e00-\u9fff]+", "", text)


def normalized_url(value: Any) -> str:
    text = str(value or "").strip().lower()
    if not text:
        return ""
    parsed = urlsplit(text)
    query = [
        (key, val) for key, val in parse_qsl(parsed.query, keep_blank_values=True)
        if key.lower() not in {"f_link_type", "flow_extra", "utm_source", "utm_medium", "utm_campaign"}
    ]
    return urlunsplit((parsed.scheme, parsed.netloc, parsed.path.rstrip("/") or "/", urlencode(query), ""))


def sql(value: Any) -> str:
    if value is None or value == "":
        return "NULL"
    return "'" + str(value).replace("\\", "\\\\").replace("'", "''") + "'"


def mysql_identifier(value: str) -> str:
    if not re.fullmatch(r"[A-Za-z0-9_]+", value):
        raise RuntimeError(f"Unsafe MySQL identifier: {value}")
    return f"`{value}`"


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest().upper()


if __name__ == "__main__":
    main()
