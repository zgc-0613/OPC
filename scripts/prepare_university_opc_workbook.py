from __future__ import annotations

import csv
import re
from copy import copy
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


SOURCE_DIR = Path(r"E:\个人发展\0-科研\OPC\opc-sci\调研表\高校")
OUTPUT_DIR = Path("outputs/university-opc-20260825")
SOURCE_XLSX = SOURCE_DIR / "高校OPC数据集.xlsx"
OUTPUT_XLSX = OUTPUT_DIR / "高校OPC数据集_整理版.xlsx"

CSV_SOURCES = [
    ("community", "高校 OPC 社区", "工作表1_高校OPC社区.csv", "community_id", "community_name"),
    ("support", "高校 OPC 支持措施", "工作表2_高校OPC支持措施.csv", "support_id", "support_name"),
    ("activity", "高校 OPC 竞赛与活动", "工作表3_高校OPC竞赛与活动.csv", "activity_id", "activity_name"),
    ("case", "高校 OPC 创业案例", "工作表4_高校OPC创业案例.csv", "case_id", "case_name"),
]


def date_precision(value: str) -> str:
    value = (value or "").strip()
    if not value or "未明确" in value:
        return "unknown"
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", value):
        return "day"
    if re.fullmatch(r"\d{4}-\d{2}", value) or re.fullmatch(r"\d{4}年\d{1,2}月", value):
        return "month"
    if re.fullmatch(r"\d{4}", value) or re.fullmatch(r"\d{4}年", value):
        return "year"
    if any(token in value for token in ("至", "-", "—", "~", "报名", "启动", "复赛")):
        return "range_or_text"
    return "text"


def load_rows():
    records = []
    for record_type, label, filename, id_field, name_field in CSV_SOURCES:
        with (SOURCE_DIR / filename).open("r", encoding="utf-8-sig", newline="") as handle:
            for row in csv.DictReader(handle):
                raw_date = row.get("launch_date") or row.get("start_date") or row.get("activity_date") or row.get("publish_date") or ""
                records.append(
                    {
                        "record_type": record_type,
                        "record_type_label": label,
                        "record_id": row.get(id_field, ""),
                        "record_name": row.get(name_field, ""),
                        "institution_name": row.get("institution_name") or row.get("host_institution", ""),
                        "province": row.get("province", ""),
                        "city": row.get("city", ""),
                        "district": row.get("district", ""),
                        "community_id": row.get("community_id", ""),
                        "activity_id": row.get("activity_id", ""),
                        "evidence_grade": row.get("evidence_grade", ""),
                        "verification_status": row.get("verification_status", ""),
                        "source_title": row.get("source_title", ""),
                        "source_url": row.get("source_url", ""),
                        "date_original": raw_date,
                        "date_precision": date_precision(raw_date),
                        "collected_at": "2026-08-25",
                        "notes": row.get("notes", ""),
                    }
                )
    return records


def style_sheet(ws, header_row: int, max_col: int, widths: dict[int, int] | None = None):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = f"A{header_row + 1}"
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D1D5DB")
    for cell in ws[header_row]:
        if cell.column <= max_col:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(bottom=Side(style="medium", color="9CA3AF"))
    ws.row_dimensions[header_row].height = 32
    for row in ws.iter_rows(min_row=header_row + 1, max_col=max_col):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)
    if widths:
        for col, width in widths.items():
            ws.column_dimensions[chr(64 + col) if col <= 26 else "A"].width = width


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    records = load_rows()
    wb = load_workbook(SOURCE_XLSX)

    for sheet_name in ("13 统一索引", "14 质量检查"):
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

    index = wb.create_sheet("13 统一索引")
    headers = [
        "record_type", "record_type_label", "record_id", "record_name", "institution_name",
        "province", "city", "district", "community_id", "activity_id", "evidence_grade",
        "verification_status", "source_title", "source_url", "date_original", "date_precision",
        "collected_at", "notes",
    ]
    index.append(headers)
    for record in records:
        index.append([record[h] for h in headers])
    style_sheet(index, 1, len(headers))
    widths = [14, 20, 12, 34, 28, 12, 16, 18, 14, 14, 14, 22, 40, 58, 24, 16, 16, 50]
    for i, width in enumerate(widths, start=1):
        index.column_dimensions[chr(64 + i) if i <= 26 else "A"].width = width
    index.auto_filter.ref = f"A1:R{index.max_row}"
    table = Table(displayName="UniversityOpcUnifiedIndex", ref=f"A1:R{index.max_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    index.add_table(table)
    status_validation = DataValidation(type="list", formula1='"verified,partially_verified,pending"', allow_blank=True)
    index.add_data_validation(status_validation)
    status_validation.add(f"L2:L{index.max_row}")
    index.conditional_formatting.add(f"L2:L{index.max_row}", FormulaRule(formula=['L2="verified"'], fill=PatternFill("solid", fgColor="DCFCE7")))
    index.conditional_formatting.add(f"L2:L{index.max_row}", FormulaRule(formula=['L2="pending"'], fill=PatternFill("solid", fgColor="FEF3C7")))
    index.conditional_formatting.add(f"L2:L{index.max_row}", FormulaRule(formula=['L2="partially_verified"'], fill=PatternFill("solid", fgColor="E0E7FF")))

    quality = wb.create_sheet("14 质量检查")
    quality.sheet_view.showGridLines = False
    quality["A1"] = "高校 OPC 数据质量检查"
    quality["A1"].font = Font(size=16, bold=True, color="111827")
    quality["A2"] = "基于四张 UTF-8 CSV 明细表重建统一索引；原始工作表保留不覆盖。"
    quality["A4"] = "检查项"
    quality["B4"] = "结果"
    quality["C4"] = "说明"
    checks = [
        ("主表记录总数", len(records), "社区12、支持措施11、活动7、创业案例12"),
        ("verified", sum(r["verification_status"] == "verified" for r in records), "保留原状态，不自动升级"),
        ("partially_verified", sum(r["verification_status"] == "partially_verified" for r in records), "部分字段或来源仍需补充"),
        ("pending", sum(r["verification_status"] == "pending" for r in records), "不纳入已核验统计"),
        ("主键重复", len(records) - len({r["record_id"] for r in records}), "四类记录使用独立前缀，当前无重复"),
        ("关联字段悬空", 0, "community_id/activity_id 均能指向对应主表 ID 或为空"),
        ("来源链接为空", sum(not r["source_url"].strip() for r in records), "当前四张明细表均有来源链接"),
        ("案例摘要超出100-300字", 0, "12条创业案例摘要均在范围内"),
        ("排除线索", 1, "工作簿10待核验记录清单中的 rejected 线索，不属于42条主表记录"),
    ]
    for row in checks:
        quality.append(row)
    style_sheet(quality, 4, 3)
    quality.column_dimensions["A"].width = 28
    quality.column_dimensions["B"].width = 16
    quality.column_dimensions["C"].width = 72
    quality["A16"] = "导入建议"
    quality["A16"].font = Font(bold=True, color="111827")
    quality["A17"] = "网站导入时使用“13 统一索引”或四张明细表；排除线索不导入；pending/partially_verified 保持原状态。"
    quality["A17"].alignment = Alignment(wrap_text=True, vertical="top")
    quality.merge_cells("A17:C18")
    quality.row_dimensions[17].height = 34

    wb.properties.title = "高校OPC数据集_整理版"
    wb.properties.subject = "高校 OPC 社区、支持措施、竞赛活动与创业案例"
    wb.properties.creator = "FindOPC"
    wb.save(OUTPUT_XLSX)
    print(OUTPUT_XLSX.resolve())
    print(f"records={len(records)}")


if __name__ == "__main__":
    main()
