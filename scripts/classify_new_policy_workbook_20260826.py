import json
from collections import Counter
from datetime import date, datetime
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
SOURCE = Path(r"e:\个人发展\0-科研\OPC\opc-sci\调研表\人工智能OPC政策汇总_新增_广东浙江新疆_0825.xlsx")
OUT = ROOT / "outputs" / "policy-new-gd-zj-xj-20260826"

LABELS = {
    "comprehensive": "综合发展政策",
    "computing_support": "算力技术",
    "funding_subsidy": "财政激励",
    "scenario_demand": "场景开放",
    "talent_service": "人才培育",
    "investment": "金融资本",
    "governance_market": "制度治理",
}

PRIMARY = {
    97: "comprehensive",
    98: "comprehensive",
    99: "governance_market",
    100: "comprehensive",
    101: "governance_market",
    102: "comprehensive",
    103: "governance_market",
    104: "investment",
    105: "funding_subsidy",
    106: "scenario_demand",
    107: "comprehensive",
    108: "comprehensive",
    109: "comprehensive",
    110: "comprehensive",
    111: "comprehensive",
    112: "governance_market",
    113: "governance_market",
    114: "comprehensive",
    115: "comprehensive",
    116: "comprehensive",
}

FIELD_TOPICS = {
    "具体政策·算力支持": "computing_support",
    "具体政策·资金补贴": "funding_subsidy",
    "具体政策·场景需求": "scenario_demand",
    "具体政策·人才服务": "talent_service",
    "具体政策·投资融资": "investment",
}

REVIEW_NOTES = {
    99: "发文单位、成文日期和文号缺失，正式入库前需从政府页面补齐。",
    100: "当前为新闻报道汇总，未见独立政府政策原文，应保持待核验。",
    104: "原文链接为商业转载，辅证为南方网；需继续查找桂城街道官方原文。",
    108: "当前链接均非政府原文，征求意见稿内容需以官方发布页复核。",
    114: "当前原文链接为文档预览页，应以政府征求意见页面替换。",
    115: "当前为中新网报道，未见政府征求意见原文，应保持待核验。",
}


def serializable(value):
    if isinstance(value, (date, datetime)):
        return value.strftime("%Y-%m-%d")
    return value


def compact(value, limit=180):
    text = " ".join(str(value or "").split())
    return text if len(text) <= limit else text[: limit - 1] + "…"


def main():
    workbook = openpyxl.load_workbook(SOURCE, read_only=True, data_only=True)
    sheet = workbook["两条修正版"]
    headers = [str(value or "").strip() for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    records = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        if not any(value not in (None, "") for value in values):
            continue
        raw = {headers[index]: serializable(value) for index, value in enumerate(values)}
        policy_id = int(raw["政策id"])
        primary = PRIMARY[policy_id]
        topics = []
        if primary == "comprehensive":
            topics.append("comprehensive")
        for field, topic in FIELD_TOPICS.items():
            if raw.get(field) not in (None, "") and topic not in topics:
                topics.append(topic)
        # Space/incubation and administrative service measures belong to the
        # entrepreneurial market environment in the closed seven-theme system.
        if raw.get("具体政策·场地工位") not in (None, "") or raw.get("具体政策·其他") not in (None, ""):
            if "governance_market" not in topics:
                topics.append("governance_market")
        if primary not in topics:
            topics.insert(0, primary)

        evidence_fields = []
        for field, topic in FIELD_TOPICS.items():
            if topic in topics and raw.get(field):
                evidence_fields.append(f"{LABELS[topic]}：{compact(raw[field])}")
        if "governance_market" in topics:
            basis = raw.get("具体政策·其他") or raw.get("具体政策·场地工位")
            if basis:
                evidence_fields.append(f"{LABELS['governance_market']}：{compact(basis)}")

        status = str(raw.get("政策状态") or "")
        review_note = REVIEW_NOTES.get(policy_id, "")
        if "征求意见稿" in status:
            review_note = ("征求意见稿，不作为现行有效政策统计。" + review_note).strip()
        records.append({
            "id": policy_id,
            "title": raw.get("title政策标题"),
            "province": raw.get("region省"),
            "city": raw.get("region市"),
            "district": raw.get("region区"),
            "status": status,
            "originalUrl": raw.get("url政策原文网页链接"),
            "evidenceUrl": raw.get("辅证链接"),
            "primaryType": primary,
            "primaryTypeLabel": LABELS[primary],
            "topicTags": topics,
            "topicTagLabels": [LABELS[topic] for topic in topics],
            "classificationEvidence": evidence_fields[:7],
            "needsAdminReview": bool(review_note),
            "reviewNote": review_note or None,
            "sourceRow": raw,
        })

    if len(records) != 20 or set(PRIMARY) != {row["id"] for row in records}:
        raise RuntimeError("record scope mismatch")

    OUT.mkdir(parents=True, exist_ok=True)
    (OUT / "classification-review.json").write_text(
        json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    lines = [
        "# 广东、浙江、新疆新增政策七类分类复核表",
        "",
        "口径：每条政策仅有一个主分类；涉及主题可多选，且只使用七类闭合词表。",
        "",
        "| ID | 政策标题 | 主分类 | 涉及主题 | 状态 | 复核提示 |",
        "|---:|---|---|---|---|---|",
    ]
    for row in records:
        lines.append(
            f"| {row['id']} | {row['title']} | {row['primaryTypeLabel']} | "
            f"{'、'.join(row['topicTagLabels'])} | {row['status']} | {row['reviewNote'] or '无'} |"
        )
    counts = Counter(row["primaryTypeLabel"] for row in records)
    lines.extend(["", "## 主分类数量", ""])
    for label in LABELS.values():
        lines.append(f"- {label}：{counts.get(label, 0)} 条")
    (OUT / "classification-review.md").write_text("\n".join(lines), encoding="utf-8")
    print(json.dumps({
        "records": len(records),
        "primaryCounts": counts,
        "adminReview": sum(row["needsAdminReview"] for row in records),
        "output": str(OUT),
    }, ensure_ascii=False, default=dict))


if __name__ == "__main__":
    main()
